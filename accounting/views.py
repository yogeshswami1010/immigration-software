from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, Count
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.urls import reverse
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
import csv
import io
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from .models import (
    Invoice,
    InvoiceItem,
    Account,
    Transaction,
    Payment,
    Expense,
    BusinessInfo,
    EmailTemplate,
    TaxSlab,
)
from rio.models import Applicant, Payments as ApplicantPayment, Services
import re


def replace_email_variables(text, invoice, business, preview_url=None):
    """
    Replace email template variables with actual values.
    Supports multiple formats: [variable], {variable}, <variable>
    """
    if not text:
        return text
    
    # Resolve "customer" fields from the Applicant model.
    # Older invoices used a dedicated Customer model; newer ones attach
    # directly to rio.Applicant via Invoice.applicant. We keep the
    # template variable names (customer_*) for backwards compatibility.
    applicant = getattr(invoice, "applicant", None)
    legacy_customer = getattr(invoice, "customer", None)

    customer_name = ""
    customer_email = ""
    customer_phone = ""

    if applicant is not None:
        name_parts = [
            getattr(applicant, "f_name", None) or "",
            getattr(applicant, "m_name", None) or "",
            getattr(applicant, "l_name", None) or "",
        ]
        full_name = " ".join(p for p in name_parts if p).strip()
        customer_name = full_name or getattr(applicant, "client_id", "") or ""
        customer_email = getattr(applicant, "email", "") or ""
        # Phone is stored as BigInteger; cast to string for display.
        phone_value = getattr(applicant, "phone", "") or ""
        customer_phone = str(phone_value) if phone_value is not None else ""
    elif legacy_customer is not None:
        # Fallback for truly legacy invoices, if any still reference a
        # Customer model instance.
        customer_name = getattr(legacy_customer, "name", "") or ""
        customer_email = getattr(legacy_customer, "email", "") or ""
        customer_phone = getattr(legacy_customer, "phone", "") or ""
    
    # Get payment-related values
    paid_amount = invoice.get_paid_amount()
    balance_due = invoice.get_balance()
    
    # Build company address
    company_address_parts = []
    if business.address_line1:
        company_address_parts.append(business.address_line1)
    if business.address_line2:
        company_address_parts.append(business.address_line2)
    city_province = []
    if business.city:
        city_province.append(business.city)
    if business.province:
        city_province.append(business.province)
    if city_province:
        company_address_parts.append(', '.join(city_province))
    if business.postal_code:
        company_address_parts.append(business.postal_code)
    if business.country:
        company_address_parts.append(business.country)
    company_address = ', '.join(company_address_parts) if company_address_parts else ''
    
    # Variable mapping
    variables = {
        'invoice_number': invoice.invoice_number,
        'customer_name': customer_name,
        'customer_email': customer_email,
        'customer_phone': customer_phone,
        'invoice_date': invoice.invoice_date.strftime('%B %d, %Y'),
        'due_date': invoice.due_date.strftime('%B %d, %Y'),
        'total_amount': f'${invoice.total_amount:.2f}',
        'subtotal': f'${invoice.subtotal:.2f}',
        'tax_amount': f'${invoice.tax_amount:.2f}',
        'discount_amount': f'${invoice.discount_amount:.2f}',
        'paid_amount': f'${paid_amount:.2f}',
        'balance_due': f'${balance_due:.2f}',
        'company_name': business.name or '',
        'company_email': business.email or '',
        'company_phone': business.phone or '',
        'company_address': company_address,
        'invoice_url': preview_url or '',
    }
    
    # Replace variables in all formats: [var], {var}, <var>
    for var_name, var_value in variables.items():
        # Square brackets: [variable]
        text = re.sub(r'\[' + re.escape(var_name) + r'\]', str(var_value), text, flags=re.IGNORECASE)
        # Curly braces: {variable}
        text = re.sub(r'\{' + re.escape(var_name) + r'\}', str(var_value), text, flags=re.IGNORECASE)
        # Angle brackets: <variable>
        text = re.sub(r'<' + re.escape(var_name) + r'>', str(var_value), text, flags=re.IGNORECASE)
    
    return text


@login_required
def invoices(request):
    """List all invoices"""
    invoices_list = Invoice.objects.all()
    
    # Filters
    invoice_number_filter = request.GET.get('invoice_number', '')
    status_filter = request.GET.get('status', '')
    # Query parameter is still called "customer" for backwards-compatibility
    # with existing templates, but it now represents an Applicant ID.
    customer_filter = request.GET.get('customer', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if invoice_number_filter:
        invoices_list = invoices_list.filter(
            invoice_number__icontains=invoice_number_filter
        )
    if status_filter:
        invoices_list = invoices_list.filter(status=status_filter)
    if customer_filter:
        invoices_list = invoices_list.filter(applicant_id=customer_filter)
    if date_from:
        invoices_list = invoices_list.filter(invoice_date__gte=date_from)
    if date_to:
        invoices_list = invoices_list.filter(invoice_date__lte=date_to)
    
    # Statistics
    total_invoiced = sum(inv.total_amount for inv in invoices_list if inv.status != 'cancelled')
    total_paid = sum(inv.get_paid_amount() for inv in invoices_list)
    total_outstanding = total_invoiced - total_paid
    
    overdue_invoices = invoices_list.filter(
        status__in=['sent', 'partial'],
        due_date__lt=timezone.now().date(),
    )
    overdue_amount = sum(inv.get_balance() for inv in overdue_invoices)

    # Build applicant choices for the "customer" filter dropdown while keeping
    # the existing template variable names.
    applicants_qs = Applicant.objects.all().order_by("f_name", "l_name", "client_id")
    customers_choices = [
        {
            "id": app.id,
            "name": (
                f"{(app.f_name or '').strip()} {(app.l_name or '').strip()}".strip()
                or app.client_id
            ),
        }
        for app in applicants_qs
    ]

    context = {
        'invoices': invoices_list,
        'customers': customers_choices,
        'invoice_number_filter': invoice_number_filter,
        'status_filter': status_filter,
        'customer_filter': customer_filter,
        'date_from': date_from,
        'date_to': date_to,
        'total_invoiced': total_invoiced,
        'total_paid': total_paid,
        'total_outstanding': total_outstanding,
        'overdue_amount': overdue_amount,
    }
    return render(request, 'accounting/invoice.html', context)


@login_required
def create_invoice(request):
    """Create a new invoice"""
    if request.method == 'POST':
        try:
            applicant_id = request.POST.get('applicant') or None
            invoice_date = request.POST.get('invoice_date')
            due_date = request.POST.get('due_date')
            notes = request.POST.get('notes', '')
            terms = request.POST.get('terms', '')

            if not applicant_id:
                messages.error(request, 'Please select an applicant for this invoice.')
                raise ValueError("Applicant is required for invoice creation")

            # Get invoice number from form or generate
            invoice_number = request.POST.get('invoice_number', '').strip()
            if not invoice_number:
                # Generate invoice number (format: RI000001)
                last_invoice = Invoice.objects.order_by('-id').first()
                if last_invoice and last_invoice.invoice_number:
                    try:
                        inv_num = last_invoice.invoice_number
                        if inv_num.upper().startswith('RI'):
                            last_num = int(inv_num[3:])
                            invoice_number = f"RI{last_num + 1:05d}"
                        elif inv_num.upper().startswith('NB'):
                            # Support legacy NB-prefixed invoices
                            last_num = int(inv_num[2:])
                            invoice_number = f"RI{last_num + 1:05d}"
                        else:
                            # Fallback for old formats like XXX-00001
                            last_num = int(inv_num.split('-')[-1])
                            invoice_number = f"RI{last_num + 1:05d}"
                    except Exception:
                        invoice_number = "RI000001"
                else:
                    invoice_number = "RI000001"

            # Check if invoice number already exists
            if Invoice.objects.filter(invoice_number=invoice_number).exists():
                messages.error(
                    request,
                    f'Invoice number {invoice_number} already exists. Please choose a different number.'
                )
                business = BusinessInfo.get_current()
                today = timezone.now().date()
                default_due_date = today + timedelta(days=business.default_due_days)
                context = {
                    'applicants': Applicant.objects.all(),
                    'tax_slabs': TaxSlab.objects.filter(is_active=True),
                    'today': today,
                    'default_due_date': default_due_date,
                    'business': business,
                }
                return render(request, 'accounting/add_invoice.html', context)

            # Create the invoice
            invoice = Invoice.objects.create(
                invoice_number=invoice_number,
                applicant_id=applicant_id,
                invoice_date=invoice_date,
                due_date=due_date,
                title=request.POST.get('title', 'Invoice'),
                summary=request.POST.get('summary', ''),
                po_number=request.POST.get('po_number', ''),
                notes=notes or '',
                terms=terms or '',
                footer=request.POST.get('footer', ''),
                created_by=request.user
            )

            # Add line items from the new applicant-based invoice form.
            # Fields are posted as:
            # - service_name[] : human-readable item/service name
            # - service[]      : optional rio.Services id
            # - quantity[]     : quantity (defaults to 1)
            # - unit_price[]   : price
            # - is_taxable[]   : "1" if this line should be included in the
            #                    taxable base when applying invoice.tax_rate.
            service_names = request.POST.getlist('service_name[]')
            service_ids = request.POST.getlist('service[]')
            quantities = request.POST.getlist('quantity[]')
            unit_prices = request.POST.getlist('unit_price[]')
            is_taxable_flags = request.POST.getlist('is_taxable[]')

            max_len = max(
                len(service_names),
                len(service_ids),
                len(quantities),
                len(unit_prices),
                len(is_taxable_flags),
            )

            for i in range(max_len):
                name = service_names[i] if i < len(service_names) else ''
                qty_raw = quantities[i] if i < len(quantities) else ''
                price_raw = unit_prices[i] if i < len(unit_prices) else ''

                # Skip completely empty rows
                if not any([name, qty_raw, price_raw]):
                    continue

                # Resolve optional service reference
                service_obj = None
                service_id = service_ids[i] if i < len(service_ids) else ''
                if service_id:
                    try:
                        service_obj = Services.objects.get(id=service_id)
                    except Services.DoesNotExist:
                        service_obj = None

                try:
                    qty = Decimal(str(qty_raw)) if qty_raw else Decimal('1.00')
                except Exception:
                    qty = Decimal('1.00')

                try:
                    unit_price = Decimal(str(price_raw)) if price_raw else Decimal('0.00')
                except Exception:
                    unit_price = Decimal('0.00')

                is_taxable = False
                if i < len(is_taxable_flags):
                    is_taxable = str(is_taxable_flags[i]).strip() == '1'

                InvoiceItem.objects.create(
                    invoice=invoice,
                    service=service_obj,
                    description=name,
                    quantity=qty,
                    unit_price=unit_price,
                    tax_rate=Decimal('0.00'),
                    is_taxable=is_taxable,
                )

            # After items are created, compute discount and tax based on posted values
            # Subtotal from items
            subtotal = sum(item.line_total for item in invoice.items.all())

            # Discount
            discount_value = request.POST.get('discount_amount', '0') or '0'
            discount_type = request.POST.get('discount_type', 'amount')
            try:
                discount_decimal = Decimal(discount_value)
            except Exception:
                discount_decimal = Decimal('0.00')

            discount_amount = Decimal('0.00')
            if discount_decimal > 0:
                if discount_type == 'percent':
                    discount_amount = (subtotal * discount_decimal) / Decimal('100.00')
                else:
                    discount_amount = discount_decimal

            # Ensure discount does not exceed subtotal
            if discount_amount > subtotal:
                discount_amount = subtotal

            invoice.discount_amount = discount_amount

            # Invoice-level tax rate from selected tax slab
            tax_slab_id = request.POST.get('tax_slab') or ''
            if tax_slab_id:
                try:
                    tax_slab = TaxSlab.objects.get(id=tax_slab_id, is_active=True)
                    invoice.tax_rate = tax_slab.rate
                except TaxSlab.DoesNotExist:
                    invoice.tax_rate = Decimal('0.00')
            else:
                invoice.tax_rate = Decimal('0.00')

            # Recalculate totals using model helper (applies discount before tax)
            invoice.calculate_totals()
            messages.success(request, 'Invoice created successfully.')
            return redirect('accounting_invoice_view', invoice_id=invoice.id)

        except Exception as e:
            messages.error(request, f'Error creating invoice: {str(e)}')
    
    business = BusinessInfo.get_current()
    # Calculate default due date based on settings
    today = timezone.now().date()
    default_due_date = today + timedelta(days=business.default_due_days)

    context = {
        'applicants': Applicant.objects.all(),
        'tax_slabs': TaxSlab.objects.filter(is_active=True),
        'today': today,
        'default_due_date': default_due_date,
        'business': business,
    }
    return render(request, 'accounting/add_invoice.html', context)


@login_required
def view_invoice(request, invoice_id):
    """View invoice details"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    # Update status automatically based on accounting payments (existing behavior)
    invoice.update_status()

    # Also compute payment summary based on Rio applicant payments. New invoices
    # attach directly to rio.Applicant via Invoice.applicant, but we keep a
    # small fallback for any legacy customer-linked records.
    applicant_payments = ApplicantPayment.objects.none()
    applicant_total_paid = Decimal('0.00')
    applicant_balance = invoice.total_amount
    applicant_payment_status = 'Unpaid'

    applicant_obj = getattr(invoice, 'applicant', None)
    if applicant_obj is None:
        # Legacy fallback: some historical invoices may still reference a
        # customer that points to an applicant.
        customer_applicant = getattr(getattr(invoice, 'customer', None), 'applicant', None)
        applicant_obj = customer_applicant

    if applicant_obj is not None:
        applicant_payments = ApplicantPayment.objects.filter(
            applicant=applicant_obj
        ).order_by('-paid_date', '-id')

        applicant_total_paid = sum(
            (Decimal(str(p.amount)) for p in applicant_payments),
            Decimal('0.00')
        )
        applicant_balance = invoice.total_amount - applicant_total_paid
        if applicant_balance < Decimal('0.00'):
            applicant_balance = Decimal('0.00')

        if applicant_total_paid <= Decimal('0.00'):
            applicant_payment_status = 'Unpaid'
        elif applicant_total_paid >= invoice.total_amount:
            applicant_payment_status = 'Paid'
        else:
            applicant_payment_status = 'Partial'

        # Reflect applicant payments in invoice status (without touching cancelled)
        if invoice.status != 'cancelled':
            new_status = invoice.status
            if applicant_payment_status == 'Paid':
                new_status = 'paid'
            elif applicant_payment_status == 'Partial':
                new_status = 'partial'

            if new_status != invoice.status:
                invoice.status = new_status
                invoice.save(update_fields=['status'])

    # Determine whether quantity column should be shown (only if any item quantity != 1)
    show_quantity = invoice.items.exclude(quantity=1).exists()

    context = {
        'invoice': invoice,
        'applicant_payments': applicant_payments,
        'applicant_total_paid': applicant_total_paid,
        'applicant_balance': applicant_balance,
        'applicant_payment_status': applicant_payment_status,
        'show_quantity': show_quantity,
    }
    return render(request, 'accounting/view_invoice.html', context)


@login_required
def change_invoice_status(request, invoice_id):
    """Change invoice status manually"""
    if request.method == 'POST':
        invoice = get_object_or_404(Invoice, id=invoice_id)
        new_status = request.POST.get('status')
        
        if new_status in dict(Invoice.STATUS_CHOICES):
            invoice.status = new_status
            invoice.save()
            messages.success(request, f'Invoice status changed to {invoice.get_status_display()}.')
        else:
            messages.error(request, 'Invalid status selected.')
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Status changed to {invoice.get_status_display()}'})
        
        return redirect('accounting_invoice_view', invoice_id=invoice.id)
    
    return redirect('accounting_invoices')


@login_required
def send_invoice(request, invoice_id):
    """Send invoice via email and change status to 'sent'"""
    from rio.services import NotificationService
    from django.urls import reverse
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    applicant = getattr(invoice, 'applicant', None)
    applicant_email = getattr(applicant, 'email', '') if applicant is not None else ''

    if not applicant_email:
        messages.error(request, 'Applicant does not have an email address. Cannot send invoice.')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Applicant email not found'}, status=400)
        return redirect('accounting_invoice_view', invoice_id=invoice.id)
    
    try:
        # Generate invoice preview URL
        preview_url = request.build_absolute_uri(reverse('accounting_invoice_preview', args=[invoice.id]))
        
        # Get business info
        business = BusinessInfo.get_current()
        
        # Get email template
        email_template = None
        original_body = None
        try:
            email_template = EmailTemplate.objects.get(template_type='invoice_sent', is_active=True)
            subject = email_template.subject
            email_body = email_template.body
            original_body = email_template.body
        except EmailTemplate.DoesNotExist:
            # Fallback to default template if no template found
            subject = f'Invoice [invoice_number] from [company_name]'
            email_body = """
            <p>Dear [customer_name],</p>
            
            <p>Please find your invoice <strong>[invoice_number]</strong> for the amount of <strong>[total_amount]</strong>.</p>
            
            <p><strong>Invoice Details:</strong></p>
            <ul>
                <li>Invoice Number: [invoice_number]</li>
                <li>Invoice Date: [invoice_date]</li>
                <li>Due Date: [due_date]</li>
                <li>Total Amount: [total_amount]</li>
            </ul>
            
            <p>You can view and download the invoice by clicking the following link:</p>
            <p><a href="[invoice_url]" style="background-color: #0d6efd; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px; display: inline-block;">View Invoice</a></p>
            
            <p>If you have any questions, please don't hesitate to contact us.</p>
            
            <p>Best regards,<br>
            [company_name]</p>
            """
            original_body = email_body
        
        # Replace variables in subject and body
        subject = replace_email_variables(subject, invoice, business, preview_url)
        email_body = replace_email_variables(email_body, invoice, business, preview_url)
        
        # Add invoice notes if available (only if not already in template)
        if invoice.notes and original_body:
            # Check if notes variable is already in template
            has_notes_var = (
                '[notes]' in original_body.lower() or 
                '{notes}' in original_body.lower() or 
                '<notes>' in original_body.lower()
            )
            
            if not has_notes_var:
                notes_section = f'<p><strong>Notes:</strong><br>{invoice.notes}</p>'
                # Insert notes before the closing message
                if '</p>' in email_body:
                    # Find the last </p> before closing tags
                    last_p = email_body.rfind('</p>', 0, email_body.rfind('</p>'))
                    if last_p > 0:
                        email_body = email_body[:last_p] + notes_section + email_body[last_p:]
                    else:
                        email_body = email_body.replace('</p>', notes_section + '</p>', 1)
        
        # Send email
        NotificationService.send_email_notification(
            subject=subject,
            body=email_body,
            recipients=[applicant_email]
        )
        
        # Change status to 'sent'
        invoice.status = 'sent'
        invoice.save()
        
        messages.success(request, f'Invoice sent successfully to {applicant_email}.')
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Invoice sent to {applicant_email}'})
        
        return redirect('accounting_invoice_view', invoice_id=invoice.id)
    
    except Exception as e:
        messages.error(request, f'Error sending invoice: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        return redirect('accounting_invoice_view', invoice_id=invoice.id)


@login_required
def edit_invoice(request, invoice_id):
    """Edit invoice"""
    invoice = get_object_or_404(Invoice, id=invoice_id)

    # Prevent editing if status is 'sent' or if any payments have been recorded
    if invoice.status == 'sent':
        messages.error(
            request,
            'This invoice has been sent and cannot be edited. Please change the status to draft first if you need to modify it.'
        )
        return redirect('accounting_invoice_view', invoice_id=invoice.id)
    
    if invoice.payments.exists():
        messages.error(
            request,
            'This invoice has recorded payments and cannot be edited. Please delete the payments first if you need to modify the invoice.'
        )
        return redirect('accounting_invoice_view', invoice_id=invoice.id)

    if request.method == 'POST':
        try:
            invoice.applicant_id = request.POST.get('applicant') or None
            invoice.invoice_date = request.POST.get('invoice_date')
            invoice.due_date = request.POST.get('due_date')
            invoice.title = request.POST.get('title', 'Invoice')
            invoice.summary = request.POST.get('summary', '')
            invoice.po_number = request.POST.get('po_number', '')
            invoice.notes = request.POST.get('notes', '')
            invoice.footer = request.POST.get('footer', '')
            invoice.status = request.POST.get('status', invoice.status)

            # Update or create line items from the applicant-based invoice form.
            item_ids = request.POST.getlist('item_id[]')
            service_names = request.POST.getlist('service_name[]')
            service_ids = request.POST.getlist('service[]')
            quantities = request.POST.getlist('quantity[]')
            unit_prices = request.POST.getlist('unit_price[]')
            is_taxable_flags = request.POST.getlist('is_taxable[]')

            # Delete removed items
            existing_ids = [int(i) for i in item_ids if i]
            InvoiceItem.objects.filter(invoice=invoice).exclude(id__in=existing_ids).delete()

            max_len = max(
                len(item_ids),
                len(service_names),
                len(service_ids),
                len(quantities),
                len(unit_prices),
                len(is_taxable_flags),
            )

            for i in range(max_len):
                item_id_raw = item_ids[i] if i < len(item_ids) else ''
                name = service_names[i] if i < len(service_names) else ''
                qty_raw = quantities[i] if i < len(quantities) else ''
                price_raw = unit_prices[i] if i < len(unit_prices) else ''

                # Skip completely empty rows
                if not any([item_id_raw, name, qty_raw, price_raw]):
                    continue

                # Resolve optional service reference
                service_obj = None
                service_id = service_ids[i] if i < len(service_ids) else ''
                if service_id:
                    try:
                        service_obj = Services.objects.get(id=service_id)
                    except Services.DoesNotExist:
                        service_obj = None

                try:
                    qty = Decimal(str(qty_raw)) if qty_raw else Decimal('1.00')
                except Exception:
                    qty = Decimal('1.00')

                try:
                    unit_price = Decimal(str(price_raw)) if price_raw else Decimal('0.00')
                except Exception:
                    unit_price = Decimal('0.00')

                is_taxable = False
                if i < len(is_taxable_flags):
                    is_taxable = str(is_taxable_flags[i]).strip() == '1'

                if item_id_raw:
                    # Update existing item
                    item = InvoiceItem.objects.get(id=int(item_id_raw), invoice=invoice)
                    item.service = service_obj
                    item.description = name
                    item.quantity = qty
                    item.unit_price = unit_price
                    item.is_taxable = is_taxable
                    # tax_rate is controlled at invoice level; keep line at 0
                    item.tax_rate = Decimal('0.00')
                    item.save()
                else:
                    # Create new item
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        service=service_obj,
                        description=name,
                        quantity=qty,
                        unit_price=unit_price,
                        tax_rate=Decimal('0.00'),
                        is_taxable=is_taxable,
                    )

            # Recalculate subtotal from items
            subtotal = sum(item.line_total for item in invoice.items.all())

            # Discount from form
            discount_value = request.POST.get('discount_amount', '0') or '0'
            discount_type = request.POST.get('discount_type', 'amount')
            try:
                discount_decimal = Decimal(discount_value)
            except Exception:
                discount_decimal = Decimal('0.00')

            discount_amount = Decimal('0.00')
            if discount_decimal > 0:
                if discount_type == 'percent':
                    discount_amount = (subtotal * discount_decimal) / Decimal('100.00')
                else:
                    discount_amount = discount_decimal

            if discount_amount > subtotal:
                discount_amount = subtotal

            invoice.discount_amount = discount_amount

            # Invoice tax rate from selected tax slab (if edit form includes it)
            tax_slab_id = request.POST.get('tax_slab') or ''
            if tax_slab_id:
                try:
                    tax_slab = TaxSlab.objects.get(id=tax_slab_id, is_active=True)
                    invoice.tax_rate = tax_slab.rate
                except TaxSlab.DoesNotExist:
                    pass

            # Persist field edits & totals
            invoice.calculate_totals()
            messages.success(request, 'Invoice updated successfully.')
            return redirect('accounting_invoice_view', invoice_id=invoice.id)
        except Exception as e:
            messages.error(request, f'Error updating invoice: {str(e)}')
    
    # Determine whether quantity column should be shown (only if any item quantity != 1)
    show_quantity = invoice.items.exclude(quantity=1).exists()

    context = {
        'invoice': invoice,
        'applicants': Applicant.objects.all(),
        'tax_slabs': TaxSlab.objects.filter(is_active=True),
        'business': BusinessInfo.get_current(),
        'show_quantity': show_quantity,
    }
    return render(request, 'accounting/edit_invoice.html', context)


@login_required
def delete_invoice(request, invoice_id):
    """Delete invoice"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    if request.method == 'POST':
        invoice.delete()
        messages.success(request, 'Invoice deleted successfully.')
        return redirect('accounting_invoices')
    return render(request, 'accounting/delete_invoice.html', {'invoice': invoice})


@login_required
def customers(request):
    """List all customers"""
    customers_list = Customer.objects.all()
    
    search = request.GET.get('search', '')
    if search:
        customers_list = customers_list.filter(
            Q(name__icontains=search) |
            Q(email__icontains=search) |
            Q(phone__icontains=search)
        )
    
    context = {
        'customers': customers_list,
        'search': search,
    }
    return render(request, 'accounting/customers.html', context)


@login_required
def create_customer(request):
    """Create a new customer"""
    if request.method == 'POST':
        try:
            applicant_id = request.POST.get('applicant', '')
            customer = Customer.objects.create(
                name=request.POST.get('name'),
                email=request.POST.get('email', ''),
                phone=request.POST.get('phone', ''),
                address=request.POST.get('address', ''),
                city=request.POST.get('city', ''),
                province=request.POST.get('province', ''),
                postal_code=request.POST.get('postal_code', ''),
                country=request.POST.get('country', 'Canada'),
                tax_id=request.POST.get('tax_id', ''),
                notes=request.POST.get('notes', ''),
                applicant_id=applicant_id if applicant_id else None,
                created_by=request.user
            )
            messages.success(request, 'Customer created successfully.')
            return redirect('accounting_customers')
        except Exception as e:
            messages.error(request, f'Error creating customer: {str(e)}')
    
    context = {
        'applicants': Applicant.objects.all(),
    }
    return render(request, 'accounting/add_customer.html', context)


@login_required
def edit_customer(request, customer_id):
    """Edit customer"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        try:
            customer.name = request.POST.get('name')
            customer.email = request.POST.get('email', '')
            customer.phone = request.POST.get('phone', '')
            customer.address = request.POST.get('address', '')
            customer.city = request.POST.get('city', '')
            customer.province = request.POST.get('province', '')
            customer.postal_code = request.POST.get('postal_code', '')
            customer.country = request.POST.get('country', 'Canada')
            customer.tax_id = request.POST.get('tax_id', '')
            customer.notes = request.POST.get('notes', '')
            customer.is_active = request.POST.get('is_active') == 'on'
            customer.applicant_id = request.POST.get('applicant', '') or None
            customer.save()
            messages.success(request, 'Customer updated successfully.')
            return redirect('accounting_customers')
        except Exception as e:
            messages.error(request, f'Error updating customer: {str(e)}')
    
    context = {
        'customer': customer,
        'applicants': Applicant.objects.all(),
    }
    return render(request, 'accounting/edit_customer.html', context)


@login_required
def products_services(request):
    """List all products and services"""
    products = ProductService.objects.all()
    
    search = request.GET.get('search', '')
    type_filter = request.GET.get('type', '')
    
    if search:
        products = products.filter(
            Q(name__icontains=search) |
            Q(sku__icontains=search) |
            Q(description__icontains=search)
        )
    if type_filter:
        products = products.filter(type=type_filter)
    
    context = {
        'products_services': products,
        'search': search,
        'type_filter': type_filter,
    }
    return render(request, 'accounting/services.html', context)


@login_required
def create_product_service(request):
    """Create a new product or service"""
    if request.method == 'POST':
        try:
            ProductService.objects.create(
                name=request.POST.get('name'),
                description=request.POST.get('description', ''),
                type=request.POST.get('type', 'service'),
                sku=request.POST.get('sku', ''),
                price=request.POST.get('price'),
                cost=request.POST.get('cost', '0.00'),
                tax_rate=request.POST.get('tax_rate', '0.00'),
                sell_this=request.POST.get('sell_this') == 'on',
                buy_this=request.POST.get('buy_this') == 'on',
                account_id=request.POST.get('account', '') or None,
            )
            messages.success(request, 'Product/Service created successfully.')
            return redirect('accounting_products_services')
        except Exception as e:
            messages.error(request, f'Error creating product/service: {str(e)}')
    
    context = {
        'accounts': Account.objects.filter(is_active=True),
    }
    return render(request, 'accounting/add_product_service.html', context)


@login_required
def edit_product_service(request, product_id):
    """Edit product or service"""
    product = get_object_or_404(ProductService, id=product_id)
    
    if request.method == 'POST':
        try:
            product.name = request.POST.get('name')
            product.description = request.POST.get('description', '')
            product.type = request.POST.get('type', 'service')
            product.sku = request.POST.get('sku', '')
            product.price = request.POST.get('price')
            product.cost = request.POST.get('cost', '0.00')
            product.tax_rate = request.POST.get('tax_rate', '0.00')
            product.account_id = request.POST.get('account', '') or None
            product.is_active = request.POST.get('is_active') == 'on'
            product.save()
            messages.success(request, 'Product/Service updated successfully.')
            return redirect('accounting_products_services')
        except Exception as e:
            messages.error(request, f'Error updating product/service: {str(e)}')
    
    context = {
        'product': product,
        'accounts': Account.objects.filter(is_active=True),
    }
    return render(request, 'accounting/edit_product_service.html', context)


@login_required
def transactions(request):
    """List all transactions"""
    transactions_list = Transaction.objects.all()
    
    type_filter = request.GET.get('type', '')
    account_filter = request.GET.get('account', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if type_filter:
        transactions_list = transactions_list.filter(transaction_type=type_filter)
    if account_filter:
        transactions_list = transactions_list.filter(account_id=account_filter)
    if date_from:
        transactions_list = transactions_list.filter(transaction_date__gte=date_from)
    if date_to:
        transactions_list = transactions_list.filter(transaction_date__lte=date_to)
    
    # Statistics
    total_deposits = transactions_list.filter(transaction_type='deposit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    total_withdrawals = transactions_list.filter(transaction_type='withdrawal').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    net_amount = total_deposits - total_withdrawals
    
    context = {
        'transactions': transactions_list,
        'accounts': Account.objects.filter(is_active=True),
        'type_filter': type_filter,
        'account_filter': account_filter,
        'date_from': date_from,
        'date_to': date_to,
        'total_deposits': total_deposits,
        'total_withdrawals': total_withdrawals,
        'net_amount': net_amount,
        'today': timezone.now().date(),
    }
    return render(request, 'accounting/transactions.html', context)


@login_required
def create_transaction(request):
    """Create a new transaction"""
    if request.method == 'POST':
        try:
            # Validate required fields
            transaction_date = request.POST.get('transaction_date')
            transaction_type = request.POST.get('transaction_type')
            account_id = request.POST.get('account')
            amount = request.POST.get('amount')
            description = request.POST.get('description')
            
            if not all([transaction_date, transaction_type, account_id, amount, description]):
                error_msg = 'All required fields must be filled.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                messages.error(request, error_msg)
                return redirect('accounting_transactions')
            
            # Validate account exists
            try:
                account = Account.objects.get(id=account_id, is_active=True)
            except Account.DoesNotExist:
                error_msg = 'Selected account does not exist or is inactive.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                messages.error(request, error_msg)
                return redirect('accounting_transactions')
            
            # Validate amount
            try:
                amount_decimal = Decimal(str(amount))
                if amount_decimal <= 0:
                    raise ValueError('Amount must be greater than zero.')
            except (ValueError, InvalidOperation) as e:
                error_msg = f'Invalid amount: {str(e)}'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                messages.error(request, error_msg)
                return redirect('accounting_transactions')
            
            transaction = Transaction.objects.create(
                transaction_date=transaction_date,
                transaction_type=transaction_type,
                account=account,
                amount=amount_decimal,
                description=description,
                reference_number=request.POST.get('reference_number', '') or None,
                category_id=request.POST.get('category', '') or None,
                created_by=request.user
            )
            messages.success(request, 'Transaction created successfully.')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Transaction created successfully.'})
            return redirect('accounting_transactions')
        except Exception as e:
            error_msg = f'Error creating transaction: {str(e)}'
            messages.error(request, error_msg)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=500)
            return redirect('accounting_transactions')
    
    context = {
        'accounts': Account.objects.filter(is_active=True),
        'today': timezone.now().date(),
    }
    return render(request, 'accounting/add_transaction.html', context)


@login_required
def delete_transaction(request, transaction_id):
    """Delete a transaction"""
    if request.method == 'POST':
        try:
            transaction = get_object_or_404(Transaction, id=transaction_id)
            transaction.delete()
            messages.success(request, 'Transaction deleted successfully.')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Transaction deleted successfully.'})
            return redirect('accounting_transactions')
        except Exception as e:
            error_msg = f'Error deleting transaction: {str(e)}'
            messages.error(request, error_msg)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=500)
            return redirect('accounting_transactions')
    
    return redirect('accounting_transactions')


@login_required
def bulk_delete_transactions(request):
    """Delete multiple transactions"""
    if request.method == 'POST':
        try:
            ids_str = request.POST.get('ids', '')
            if not ids_str:
                return JsonResponse({'success': False, 'message': 'No transactions selected.'}, status=400)
            
            ids = [int(id.strip()) for id in ids_str.split(',') if id.strip().isdigit()]
            if not ids:
                return JsonResponse({'success': False, 'message': 'Invalid transaction IDs.'}, status=400)
            
            transactions = Transaction.objects.filter(id__in=ids)
            count = transactions.count()
            transactions.delete()
            
            message = f'{count} transaction(s) deleted successfully.'
            messages.success(request, message)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': message})
            return redirect('accounting_transactions')
        except Exception as e:
            error_msg = f'Error deleting transactions: {str(e)}'
            messages.error(request, error_msg)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=500)
            return redirect('accounting_transactions')
    
    return redirect('accounting_transactions')


@login_required
def bulk_categorize_transactions(request):
    """Categorize multiple transactions"""
    if request.method == 'POST':
        try:
            ids_str = request.POST.get('ids', '')
            category_id = request.POST.get('category_id', '') or None
            
            if not ids_str:
                return JsonResponse({'success': False, 'message': 'No transactions selected.'}, status=400)
            
            ids = [int(id.strip()) for id in ids_str.split(',') if id.strip().isdigit()]
            if not ids:
                return JsonResponse({'success': False, 'message': 'Invalid transaction IDs.'}, status=400)
            
            transactions = Transaction.objects.filter(id__in=ids)
            count = transactions.count()
            
            if category_id:
                try:
                    category = Account.objects.get(id=category_id, is_active=True)
                    transactions.update(category=category)
                except Account.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Invalid category selected.'}, status=400)
            else:
                transactions.update(category=None)
            
            message = f'{count} transaction(s) categorized successfully.'
            messages.success(request, message)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': message})
            return redirect('accounting_transactions')
        except Exception as e:
            error_msg = f'Error categorizing transactions: {str(e)}'
            messages.error(request, error_msg)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=500)
            return redirect('accounting_transactions')
    
    return redirect('accounting_transactions')


@login_required
def chart_of_accounts(request):
    """List chart of accounts"""
    accounts = Account.objects.filter(is_active=True)
    
    type_filter = request.GET.get('type', '')
    if type_filter:
        accounts = accounts.filter(account_type=type_filter)
    
    # Get counts for each account type
    asset_count = Account.objects.filter(account_type='asset', is_active=True).count()
    liability_count = Account.objects.filter(account_type='liability', is_active=True).count()
    revenue_count = Account.objects.filter(account_type='revenue', is_active=True).count()
    expense_count = Account.objects.filter(account_type='expense', is_active=True).count()
    equity_count = Account.objects.filter(account_type='equity', is_active=True).count()
    
    context = {
        'accounts': accounts,
        'type_filter': type_filter,
        'asset_count': asset_count,
        'liability_count': liability_count,
        'revenue_count': revenue_count,
        'expense_count': expense_count,
        'equity_count': equity_count,
    }
    return render(request, 'accounting/chart_of_accounts.html', context)


@login_required
def create_account(request):
    """Create a new account"""
    if request.method == 'POST':
        try:
            Account.objects.create(
                name=request.POST.get('name'),
                account_type=request.POST.get('account_type'),
                account_number=request.POST.get('account_number', ''),
                description=request.POST.get('description', ''),
                parent_account_id=request.POST.get('parent_account', '') or None,
            )
            messages.success(request, 'Account created successfully.')
            return redirect('accounting_chart_of_accounts')
        except Exception as e:
            messages.error(request, f'Error creating account: {str(e)}')
    
    # Pre-select account type from URL parameter
    account_type = request.GET.get('type', '')
    category = request.GET.get('category', '')
    
    context = {
        'accounts': Account.objects.filter(is_active=True),
        'account_type': account_type,
        'category': category,
    }
    return render(request, 'accounting/add_account.html', context)


@login_required
def edit_account(request, account_id):
    """Edit account"""
    account = get_object_or_404(Account, id=account_id)
    
    if request.method == 'POST':
        try:
            account.name = request.POST.get('name')
            account.account_type = request.POST.get('account_type')
            account.account_number = request.POST.get('account_number', '')
            account.description = request.POST.get('description', '')
            account.parent_account_id = request.POST.get('parent_account', '') or None
            account.is_active = request.POST.get('is_active') == 'on'
            account.save()
            messages.success(request, 'Account updated successfully.')
            return redirect('accounting_chart_of_accounts')
        except Exception as e:
            messages.error(request, f'Error updating account: {str(e)}')
    
    context = {
        'account': account,
        'accounts': Account.objects.filter(is_active=True).exclude(id=account_id),
    }
    return render(request, 'accounting/edit_account.html', context)


@login_required
def expenses(request):
    """List all expenses"""
    expenses_list = Expense.objects.all()
    
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if status_filter:
        expenses_list = expenses_list.filter(payment_status=status_filter)
    if date_from:
        expenses_list = expenses_list.filter(expense_date__gte=date_from)
    if date_to:
        expenses_list = expenses_list.filter(expense_date__lte=date_to)
    
    total_expenses = expenses_list.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    
    context = {
        'expenses': expenses_list,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'total_expenses': total_expenses,
    }
    return render(request, 'accounting/expenses.html', context)


@login_required
def create_expense(request):
    """Create a new expense"""
    if request.method == 'POST':
        try:
            expense = Expense.objects.create(
                expense_date=request.POST.get('expense_date'),
                vendor=request.POST.get('vendor'),
                description=request.POST.get('description'),
                amount=request.POST.get('amount'),
                category_id=request.POST.get('category', '') or None,
                payment_method=request.POST.get('payment_method', 'cash'),
                payment_status=request.POST.get('payment_status', 'unpaid'),
                reference_number=request.POST.get('reference_number', ''),
                account_id=request.POST.get('account', '') or None,
                created_by=request.user
            )
            if 'receipt' in request.FILES:
                expense.receipt = request.FILES['receipt']
                expense.save()
            messages.success(request, 'Expense created successfully.')
            return redirect('accounting_expenses')
        except Exception as e:
            messages.error(request, f'Error creating expense: {str(e)}')
    
    context = {
        'accounts': Account.objects.filter(account_type='expense', is_active=True),
        'today': timezone.now().date(),
    }
    return render(request, 'accounting/add_expense.html', context)


@login_required
def payments(request):
    """List all payments"""
    payments_list = Payment.objects.select_related('applicant', 'invoice').all()

    # Filter parameter name kept as "customer" for template compatibility;
    # it now represents an Applicant ID.
    customer_filter = request.GET.get('customer', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if customer_filter:
        payments_list = payments_list.filter(applicant_id=customer_filter)
    if date_from:
        payments_list = payments_list.filter(payment_date__gte=date_from)
    if date_to:
        payments_list = payments_list.filter(payment_date__lte=date_to)
    
    total_payments = payments_list.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    
    # Build applicant choices for the filter dropdown but keep the variable
    # name "customers" so that the existing template continues to work.
    applicants_qs = Applicant.objects.all().order_by('f_name', 'l_name', 'client_id')
    customers_choices = [
        {
            'id': app.id,
            'name': (
                f"{(app.f_name or '').strip()} {(app.l_name or '').strip()}".strip()
                or app.client_id
            ),
        }
        for app in applicants_qs
    ]

    context = {
        'payments': payments_list,
        'customers': customers_choices,
        'customer_filter': customer_filter,
        'date_from': date_from,
        'date_to': date_to,
        'total_payments': total_payments,
    }
    return render(request, 'accounting/payments.html', context)


@login_required
def create_payment(request):
    """Create a new payment"""
    if request.method == 'POST':
        try:
            invoice_id = request.POST.get('invoice', '')
            payment = Payment.objects.create(
                invoice_id=invoice_id if invoice_id else None,
                applicant_id=request.POST.get('customer') or None,
                payment_date=request.POST.get('payment_date'),
                amount=request.POST.get('amount'),
                payment_method=request.POST.get('payment_method', 'cash'),
                reference_number=request.POST.get('reference_number', ''),
                notes=request.POST.get('notes', ''),
                account_id=request.POST.get('account', '') or None,
                created_by=request.user
            )
            
            # Create transaction if account is specified
            if payment.account:
                applicant = payment.applicant
                if applicant is not None:
                    name_label = (
                        f"{(applicant.f_name or '').strip()} {(applicant.l_name or '').strip()}".strip()
                        or applicant.client_id
                    )
                else:
                    name_label = "Unknown applicant"
                Transaction.objects.create(
                    transaction_date=payment.payment_date,
                    transaction_type='deposit',
                    account=payment.account,
                    amount=payment.amount,
                    description=f"Payment from {name_label}",
                    reference_number=payment.reference_number,
                    payment=payment,
                    created_by=request.user
                )
            
            # Update invoice status automatically
            if payment.invoice:
                payment.invoice.update_status()
            
            messages.success(request, 'Payment created successfully.')
            return redirect('accounting_payments')
        except Exception as e:
            messages.error(request, f'Error creating payment: {str(e)}')
    
    # Get invoice from query parameter if provided
    invoice_id = request.GET.get('invoice', '')
    selected_invoice = None
    invoices = Invoice.objects.filter(status__in=['sent', 'partial'])
    
    if invoice_id:
        try:
            selected_invoice = Invoice.objects.get(id=invoice_id)
            # Include the selected invoice in the list even if not in sent/partial status
            if selected_invoice.status not in ['sent', 'partial']:
                # Use union to combine querysets
                from django.db.models import Q
                invoices = Invoice.objects.filter(Q(status__in=['sent', 'partial']) | Q(id=selected_invoice.id))
        except Invoice.DoesNotExist:
            pass
    
    # Reuse applicant-as-customer mapping for the add-payment form.
    applicants_qs = Applicant.objects.all().order_by('f_name', 'l_name', 'client_id')
    customers_choices = [
        {
            'id': app.id,
            'name': (
                f"{(app.f_name or '').strip()} {(app.l_name or '').strip()}".strip()
                or app.client_id
            ),
        }
        for app in applicants_qs
    ]

    context = {
        'customers': customers_choices,
        'invoices': invoices,
        'accounts': Account.objects.filter(account_type__in=['asset', 'revenue'], is_active=True),
        'today': timezone.now().date(),
        'selected_invoice': selected_invoice,
    }
    return render(request, 'accounting/add_payment.html', context)


@login_required
def customer_statements(request):
    """Customer statements"""
    customer_id = request.GET.get('customer', '')
    statement_type = request.GET.get('type', 'outstanding')
    
    customer = None
    invoices = Invoice.objects.none()
    overdue_amount = Decimal('0.00')
    not_due_amount = Decimal('0.00')
    today = timezone.now().date()
    
    if customer_id:
        customer = get_object_or_404(Customer, id=customer_id)
        if statement_type == 'outstanding':
            invoices = customer.invoices.filter(status__in=['sent', 'partial', 'overdue'])
        else:
            invoices = customer.invoices.all()
        
        # Calculate overdue and not due amounts
        for invoice in invoices:
            balance = invoice.get_balance()
            if balance > 0:
                if invoice.due_date < today:
                    overdue_amount += balance
                else:
                    not_due_amount += balance
    
    context = {
        'customer': customer,
        'customers': Customer.objects.filter(is_active=True),
        'invoices': invoices,
        'customer_id': customer_id,
        'overdue_amount': overdue_amount,
        'not_due_amount': not_due_amount,
        'today': today,
    }
    return render(request, 'accounting/customer_statements.html', context)


@login_required
def reports(request):
    """Comprehensive Reports Dashboard"""
    report_type = request.GET.get('report_type', '')
    export_format = request.GET.get('export', '')
    
    # Get filter parameters
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))
    account_id = request.GET.get('account', '')
    customer_id = request.GET.get('customer', '')
    transaction_type = request.GET.get('transaction_type', '')
    invoice_status = request.GET.get('invoice_status', '')
    expense_status = request.GET.get('expense_status', '')
    
    # If export is requested, generate export file
    if export_format and report_type:
        return generate_report_export(request, report_type, export_format, {
            'date_from': date_from,
            'date_to': date_to,
            'account_id': account_id,
            'customer_id': customer_id,
            'transaction_type': transaction_type,
            'invoice_status': invoice_status,
            'expense_status': expense_status,
        })
    
    # Prepare context for reports
    context = {
        'report_type': report_type,
        'date_from': date_from,
        'date_to': date_to,
        'account_id': account_id,
        'customer_id': customer_id,
        'transaction_type': transaction_type,
        'invoice_status': invoice_status,
        'expense_status': expense_status,
        'accounts': Account.objects.filter(is_active=True),
        'customers': customer.objects.filter(is_active=True),
        'OPENPYXL_AVAILABLE': OPENPYXL_AVAILABLE,
        'REPORTLAB_AVAILABLE': REPORTLAB_AVAILABLE,
    }
    
    # Generate report data based on type
    if report_type:
        context.update(generate_report_data(report_type, {
            'date_from': date_from,
            'date_to': date_to,
            'account_id': account_id,
            'customer_id': customer_id,
            'transaction_type': transaction_type,
            'invoice_status': invoice_status,
            'expense_status': expense_status,
        }))
    
    return render(request, 'accounting/reports.html', context)


def generate_report_data(report_type, filters):
    """Generate data for different report types"""
    data = {}
    date_from = filters.get('date_from')
    date_to = filters.get('date_to')
    account_id = filters.get('account_id')
    customer_id = filters.get('customer_id')
    transaction_type = filters.get('transaction_type')
    invoice_status = filters.get('invoice_status')
    expense_status = filters.get('expense_status')
    
    if report_type == 'profit_loss':
        # Revenue
        revenue_query = InvoiceItem.objects.filter(
            invoice__invoice_date__gte=date_from,
            invoice__invoice_date__lte=date_to,
            invoice__status__in=['sent', 'paid', 'partial']
        )
        if customer_id:
            revenue_query = revenue_query.filter(invoice__customer_id=customer_id)
        
        revenue_items = revenue_query.select_related('invoice', 'product_service')
        total_revenue = sum(item.line_total for item in revenue_items)
        
        # Expenses
        expense_query = Expense.objects.filter(
            expense_date__gte=date_from,
            expense_date__lte=date_to
        )
        if account_id:
            expense_query = expense_query.filter(account_id=account_id)
        if expense_status:
            expense_query = expense_query.filter(payment_status=expense_status)
        
        expenses = expense_query
        total_expenses = sum(exp.amount for exp in expenses)
        net_profit = total_revenue - total_expenses
        
        data = {
            'revenue_items': revenue_items,
            'expenses': expenses,
            'total_revenue': total_revenue,
            'total_expenses': total_expenses,
            'net_profit': net_profit,
        }
    
    elif report_type == 'transactions':
        query = Transaction.objects.select_related('account', 'category').filter(
            transaction_date__gte=date_from,
            transaction_date__lte=date_to
        )
        if account_id:
            query = query.filter(account_id=account_id)
        if transaction_type:
            query = query.filter(transaction_type=transaction_type)
        
        transactions = query.order_by('-transaction_date')
        total_deposits = sum(t.amount for t in transactions if t.transaction_type == 'deposit')
        total_withdrawals = sum(t.amount for t in transactions if t.transaction_type == 'withdrawal')
        
        data = {
            'transactions': transactions,
            'total_deposits': total_deposits,
            'total_withdrawals': total_withdrawals,
            'net_amount': total_deposits - total_withdrawals,
        }
    
    elif report_type == 'invoices':
        query = Invoice.objects.select_related('customer').filter(
            invoice_date__gte=date_from,
            invoice_date__lte=date_to
        )
        if customer_id:
            query = query.filter(customer_id=customer_id)
        if invoice_status:
            query = query.filter(status=invoice_status)
        
        invoices = query.order_by('-invoice_date')
        total_amount = sum(inv.total_amount for inv in invoices)
        total_paid = sum(inv.get_paid_amount() for inv in invoices)
        
        data = {
            'invoices': invoices,
            'total_amount': total_amount,
            'total_paid': total_paid,
            'total_outstanding': total_amount - total_paid,
        }
    
    elif report_type == 'expenses':
        query = Expense.objects.select_related('category', 'account').filter(
            expense_date__gte=date_from,
            expense_date__lte=date_to
        )
        if account_id:
            query = query.filter(account_id=account_id)
        if expense_status:
            query = query.filter(payment_status=expense_status)
        
        expenses = query.order_by('-expense_date')
        total_amount = sum(exp.amount for exp in expenses)
        
        data = {
            'expenses': expenses,
            'total_amount': total_amount,
        }
    
    elif report_type == 'payments':
        query = Payment.objects.select_related('customer', 'invoice', 'account').filter(
            payment_date__gte=date_from,
            payment_date__lte=date_to
        )
        if customer_id:
            query = query.filter(customer_id=customer_id)
        if account_id:
            query = query.filter(account_id=account_id)
        
        payments = query.order_by('-payment_date')
        total_amount = sum(pay.amount for pay in payments)
        
        data = {
            'payments': payments,
            'total_amount': total_amount,
        }
    
    elif report_type == 'balance_sheet':
        as_of_date = date_to
        asset_accounts = Account.objects.filter(account_type='asset', is_active=True)
        liability_accounts = Account.objects.filter(account_type='liability', is_active=True)
        equity_accounts = Account.objects.filter(account_type='equity', is_active=True)
        
        total_assets = sum(acc.get_balance() for acc in asset_accounts)
        total_liabilities = sum(acc.get_balance() for acc in liability_accounts)
        total_equity = sum(acc.get_balance() for acc in equity_accounts)
        
        data = {
            'as_of_date': as_of_date,
            'asset_accounts': asset_accounts,
            'liability_accounts': liability_accounts,
            'equity_accounts': equity_accounts,
            'total_assets': total_assets,
            'total_liabilities': total_liabilities,
            'total_equity': total_equity,
        }
    
    return data


def generate_report_export(request, report_type, export_format, filters):
    """Generate export file for reports"""
    data = generate_report_data(report_type, filters)
    
    if export_format == 'csv':
        return export_to_csv(report_type, data, filters)
    elif export_format == 'excel' and OPENPYXL_AVAILABLE:
        return export_to_excel(report_type, data, filters)
    elif export_format == 'pdf' and REPORTLAB_AVAILABLE:
        return export_to_pdf(report_type, data, filters)
    else:
        messages.error(request, f'{export_format.upper()} export is not available. Please install required libraries.')
        return redirect('accounting_reports')


def export_to_csv(report_type, data, filters):
    """Export report data to CSV"""
    response = HttpResponse(content_type='text/csv')
    filename = f'{report_type}_report_{filters.get("date_from")}_to_{filters.get("date_to")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    if report_type == 'profit_loss':
        writer.writerow(['Profit & Loss Report'])
        writer.writerow([f'Period: {filters.get("date_from")} to {filters.get("date_to")}'])
        writer.writerow([])
        writer.writerow(['Revenue'])
        writer.writerow(['Invoice', 'Date', 'Customer', 'Product/Service', 'Amount'])
        for item in data.get('revenue_items', []):
            writer.writerow([
                item.invoice.invoice_number,
                item.invoice.invoice_date.strftime('%Y-%m-%d'),
                item.invoice.customer.name if item.invoice.customer else '',
                item.product_service.name if item.product_service else '',
                f'{item.line_total:.2f}'
            ])
        writer.writerow(['Total Revenue', '', '', '', f'{data.get("total_revenue", 0):.2f}'])
        writer.writerow([])
        writer.writerow(['Expenses'])
        writer.writerow(['Date', 'Vendor', 'Description', 'Amount', 'Status'])
        for exp in data.get('expenses', []):
            writer.writerow([
                exp.expense_date.strftime('%Y-%m-%d'),
                exp.vendor,
                exp.description,
                f'{exp.amount:.2f}',
                exp.get_payment_status_display()
            ])
        writer.writerow(['Total Expenses', '', '', '', f'{data.get("total_expenses", 0):.2f}'])
        writer.writerow([])
        writer.writerow(['Net Profit', '', '', '', f'{data.get("net_profit", 0):.2f}'])
    
    elif report_type == 'transactions':
        writer.writerow(['Transactions Report'])
        writer.writerow([f'Period: {filters.get("date_from")} to {filters.get("date_to")}'])
        writer.writerow([])
        writer.writerow(['Date', 'Type', 'Account', 'Description', 'Amount', 'Category'])
        for trans in data.get('transactions', []):
            writer.writerow([
                trans.transaction_date.strftime('%Y-%m-%d'),
                trans.get_transaction_type_display(),
                trans.account.name,
                trans.description,
                f'{trans.amount:.2f}',
                trans.category.name if trans.category else 'Uncategorized'
            ])
        writer.writerow([])
        writer.writerow(['Total Deposits', '', '', '', f'{data.get("total_deposits", 0):.2f}'])
        writer.writerow(['Total Withdrawals', '', '', '', f'{data.get("total_withdrawals", 0):.2f}'])
        writer.writerow(['Net Amount', '', '', '', f'{data.get("net_amount", 0):.2f}'])
    
    elif report_type == 'invoices':
        writer.writerow(['Invoices Report'])
        writer.writerow([f'Period: {filters.get("date_from")} to {filters.get("date_to")}'])
        writer.writerow([])
        writer.writerow(['Invoice #', 'Date', 'Customer', 'Amount', 'Paid', 'Outstanding', 'Status'])
        for inv in data.get('invoices', []):
            writer.writerow([
                inv.invoice_number,
                inv.invoice_date.strftime('%Y-%m-%d'),
                inv.customer.name if inv.customer else '',
                f'{inv.total_amount:.2f}',
                f'{inv.get_paid_amount():.2f}',
                f'{(inv.total_amount - inv.get_paid_amount()):.2f}',
                inv.get_status_display()
            ])
        writer.writerow([])
        writer.writerow(['Total Amount', '', '', f'{data.get("total_amount", 0):.2f}'])
        writer.writerow(['Total Paid', '', '', f'{data.get("total_paid", 0):.2f}'])
        writer.writerow(['Total Outstanding', '', '', f'{data.get("total_outstanding", 0):.2f}'])
    
    elif report_type == 'expenses':
        writer.writerow(['Expenses Report'])
        writer.writerow([f'Period: {filters.get("date_from")} to {filters.get("date_to")}'])
        writer.writerow([])
        writer.writerow(['Date', 'Vendor', 'Description', 'Amount', 'Category', 'Status'])
        for exp in data.get('expenses', []):
            writer.writerow([
                exp.expense_date.strftime('%Y-%m-%d'),
                exp.vendor,
                exp.description,
                f'{exp.amount:.2f}',
                exp.category.name if exp.category else '',
                exp.get_payment_status_display()
            ])
        writer.writerow([])
        writer.writerow(['Total Amount', '', '', f'{data.get("total_amount", 0):.2f}'])
    
    elif report_type == 'payments':
        writer.writerow(['Payments Report'])
        writer.writerow([f'Period: {filters.get("date_from")} to {filters.get("date_to")}'])
        writer.writerow([])
        writer.writerow(['Date', 'Customer', 'Invoice', 'Amount', 'Method', 'Account'])
        for pay in data.get('payments', []):
            writer.writerow([
                pay.payment_date.strftime('%Y-%m-%d'),
                pay.customer.name if pay.customer else '',
                pay.invoice.invoice_number if pay.invoice else '',
                f'{pay.amount:.2f}',
                pay.get_payment_method_display(),
                pay.account.name if pay.account else ''
            ])
        writer.writerow([])
        writer.writerow(['Total Amount', '', '', f'{data.get("total_amount", 0):.2f}'])
    
    return response


def export_to_excel(report_type, data, filters):
    """Export report data to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.replace('_', ' ').title()
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    if report_type == 'profit_loss':
        ws['A1'] = 'Profit & Loss Report'
        ws['A2'] = f'Period: {filters.get("date_from")} to {filters.get("date_to")}'
        row = 4
        ws.cell(row, 1, 'Revenue')
        row += 1
        headers = ['Invoice', 'Date', 'Customer', 'Product/Service', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1
        for item in data.get('revenue_items', []):
            ws.cell(row, 1, item.invoice.invoice_number)
            ws.cell(row, 2, item.invoice.invoice_date.strftime('%Y-%m-%d'))
            ws.cell(row, 3, item.invoice.customer.name if item.invoice.customer else '')
            ws.cell(row, 4, item.product_service.name if item.product_service else '')
            ws.cell(row, 5, float(item.line_total))
            row += 1
        ws.cell(row, 4, 'Total Revenue')
        ws.cell(row, 5, float(data.get('total_revenue', 0)))
        row += 2
        ws.cell(row, 1, 'Expenses')
        row += 1
        headers = ['Date', 'Vendor', 'Description', 'Amount', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1
        for exp in data.get('expenses', []):
            ws.cell(row, 1, exp.expense_date.strftime('%Y-%m-%d'))
            ws.cell(row, 2, exp.vendor)
            ws.cell(row, 3, exp.description)
            ws.cell(row, 4, float(exp.amount))
            ws.cell(row, 5, exp.get_payment_status_display())
            row += 1
        ws.cell(row, 3, 'Total Expenses')
        ws.cell(row, 4, float(data.get('total_expenses', 0)))
        row += 2
        ws.cell(row, 3, 'Net Profit')
        ws.cell(row, 4, float(data.get('net_profit', 0)))
    
    # Similar implementation for other report types...
    # (I'll add a simplified version for now)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{report_type}_report_{filters.get("date_from")}_to_{filters.get("date_to")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_to_pdf(report_type, data, filters):
    """Export report data to PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f'{report_type.replace("_", " ").title()} Report', styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Period
    period = Paragraph(f'Period: {filters.get("date_from")} to {filters.get("date_to")}', styles['Normal'])
    elements.append(period)
    elements.append(Spacer(1, 12))
    
    # Data table (simplified - would need full implementation for each report type)
    if report_type == 'profit_loss':
        table_data = [['Revenue', f'${data.get("total_revenue", 0):.2f}'],
                      ['Expenses', f'${data.get("total_expenses", 0):.2f}'],
                      ['Net Profit', f'${data.get("net_profit", 0):.2f}']]
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f'{report_type}_report_{filters.get("date_from")}_to_{filters.get("date_to")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def profit_loss_report(request):
    """Profit & Loss Report"""
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=365)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))
    
    # Revenue
    revenue_accounts = Account.objects.filter(account_type='revenue', is_active=True)
    revenue_items = InvoiceItem.objects.filter(
        invoice__invoice_date__gte=date_from,
        invoice__invoice_date__lte=date_to,
        invoice__status__in=['sent', 'paid', 'partial']
    )
    total_revenue = sum(item.line_total for item in revenue_items)
    
    # Expenses
    expense_accounts = Account.objects.filter(account_type='expense', is_active=True)
    expenses = Expense.objects.filter(
        expense_date__gte=date_from,
        expense_date__lte=date_to
    )
    total_expenses = sum(exp.amount for exp in expenses)
    
    net_profit = total_revenue - total_expenses
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'revenue_accounts': revenue_accounts,
        'expense_accounts': expense_accounts,
    }
    return render(request, 'accounting/reports/profit_loss.html', context)


@login_required
def balance_sheet_report(request):
    """Balance Sheet Report"""
    as_of_date = request.GET.get('as_of_date', timezone.now().strftime('%Y-%m-%d'))
    
    # Assets
    asset_accounts = Account.objects.filter(account_type='asset', is_active=True)
    
    # Liabilities
    liability_accounts = Account.objects.filter(account_type='liability', is_active=True)
    
    # Equity
    equity_accounts = Account.objects.filter(account_type='equity', is_active=True)
    
    context = {
        'as_of_date': as_of_date,
        'asset_accounts': asset_accounts,
        'liability_accounts': liability_accounts,
        'equity_accounts': equity_accounts,
    }
    return render(request, 'accounting/reports/balance_sheet.html', context)


@login_required
def get_product_service_details(request, product_id):
    """AJAX endpoint to get product/service details"""
    product = get_object_or_404(ProductService, id=product_id)
    return JsonResponse({
        'price': str(product.price),
        'tax_rate': str(product.tax_rate),
        'description': product.description or '',
    })


@login_required
def get_customer_invoices(request, customer_id):
    """AJAX endpoint to get customer's unpaid invoices"""
    customer = get_object_or_404(Customer, id=customer_id)
    invoices = customer.invoices.filter(status__in=['sent', 'partial'])
    return JsonResponse({
        'invoices': [
            {
                'id': inv.id,
                'invoice_number': inv.invoice_number,
                'total_amount': str(inv.total_amount),
                'balance': str(inv.get_balance()),
            }
            for inv in invoices
        ]
    })


@login_required
def get_invoice_number(request):
    """AJAX endpoint to generate invoice number"""
    last_invoice = Invoice.objects.order_by('-id').first()
    if last_invoice:
        try:
            inv_num = last_invoice.invoice_number
            if inv_num.upper().startswith('RI'):
                last_num = int(inv_num[3:])
                invoice_number = f"RI{last_num + 1:05d}"
            elif inv_num.upper().startswith('NB'):
                # Support legacy NB-prefixed invoices
                last_num = int(inv_num[2:])
                invoice_number = f"RI{last_num + 1:05d}"
            else:
                last_num = int(inv_num.split('-')[-1])
                invoice_number = f"RI{last_num + 1:05d}"
        except Exception:
            invoice_number = "RI000001"
    else:
        invoice_number = "RI00001"
    
    return JsonResponse({'invoice_number': invoice_number})


@login_required
def business_info(request):
    """Update business information"""
    business = BusinessInfo.get_current()
    
    if request.method == 'POST':
        try:
            business.name = request.POST.get('name')
            business.address_line1 = request.POST.get('address_line1')
            business.address_line2 = request.POST.get('address_line2', '')
            business.city = request.POST.get('city')
            business.province = request.POST.get('province')
            business.postal_code = request.POST.get('postal_code')
            business.country = request.POST.get('country', 'Canada')
            business.phone = request.POST.get('phone')
            business.mobile = request.POST.get('mobile')
            business.email = request.POST.get('email', '')
            business.website = request.POST.get('website', '')
            business.tax_id = request.POST.get('tax_id', '')
            
            if 'logo' in request.FILES:
                business.logo = request.FILES['logo']
            
            business.save()
            messages.success(request, 'Business information updated successfully.')
            return redirect('accounting_business_info')
        except Exception as e:
            messages.error(request, f'Error updating business information: {str(e)}')
    
    context = {
        'business': business,
    }
    return render(request, 'accounting/business_info.html', context)


@login_required
def invoice_settings(request):
    """Comprehensive invoice settings page with tabs"""
    business = BusinessInfo.get_current()
    
    # Initialize default email templates if they don't exist
    template_types = ['invoice_sent', 'invoice_reminder', 'payment_received']
    for template_type in template_types:
        EmailTemplate.objects.get_or_create(
            template_type=template_type,
            defaults={
                'subject': f'Invoice {template_type.replace("_", " ").title()}',
                'body': f'Default template for {template_type.replace("_", " ")}',
            }
        )
    
    templates = EmailTemplate.objects.all()
    tax_slabs = TaxSlab.objects.all()
    
    if request.method == 'POST':
        try:
            tab = request.POST.get('tab', 'company')
            
            # Company Details Tab
            if tab == 'company':
                business.name = request.POST.get('name')
                business.address_line1 = request.POST.get('address_line1')
                business.address_line2 = request.POST.get('address_line2', '')
                business.city = request.POST.get('city')
                business.province = request.POST.get('province')
                business.postal_code = request.POST.get('postal_code')
                business.country = request.POST.get('country', 'Canada')
                business.phone = request.POST.get('phone')
                business.mobile = request.POST.get('mobile')
                business.email = request.POST.get('email', '')
                business.website = request.POST.get('website', '')
                business.tax_id = request.POST.get('tax_id', '')
            
            # Email Templates Tab
            elif tab == 'email':
                for template in templates:
                    template_id = request.POST.get(f'template_id_{template.id}')
                    if template_id:
                        template.subject = request.POST.get(f'subject_{template.id}', '')
                        template.body = request.POST.get(f'body_{template.id}', '')
                        template.is_active = request.POST.get(f'is_active_{template.id}') == 'on'
                        template.save()
            
            # Tax Slabs Tab
            elif tab == 'tax':
                # Update existing tax slabs
                for tax in tax_slabs:
                    tax_name = request.POST.get(f'tax_name_{tax.id}')
                    tax_rate = request.POST.get(f'tax_rate_{tax.id}')
                    if tax_name and tax_rate:
                        tax.name = tax_name
                        tax.rate = Decimal(tax_rate)
                        tax.is_default = request.POST.get(f'tax_default_{tax.id}') == 'on'
                        tax.is_active = request.POST.get(f'tax_active_{tax.id}') == 'on'
                        tax.save()
                
                # Create new tax slabs
                for key in request.POST.keys():
                    if key.startswith('tax_name_new_'):
                        counter = key.replace('tax_name_new_', '')
                        tax_name = request.POST.get(f'tax_name_new_{counter}')
                        tax_rate = request.POST.get(f'tax_rate_new_{counter}')
                        if tax_name and tax_rate:
                            TaxSlab.objects.create(
                                name=tax_name,
                                rate=Decimal(tax_rate),
                                is_default=request.POST.get(f'tax_default_new_{counter}') == 'on',
                                is_active=request.POST.get(f'tax_active_new_{counter}') == 'on'
                            )
                
                # If a tax is set as default, unset others
                default_tax = TaxSlab.objects.filter(is_default=True).first()
                if default_tax:
                    TaxSlab.objects.exclude(id=default_tax.id).update(is_default=False)
            
            # Footer Tab
            elif tab == 'footer':
                business.default_footer = request.POST.get('default_footer', '')
            
            # Notes / Terms Tab
            elif tab == 'notes':
                business.default_notes_terms = request.POST.get('default_notes_terms', '')
            
            # Invoice Settings Tab
            elif tab == 'invoice':
                default_due_days = request.POST.get('default_due_days')
                if default_due_days:
                    business.default_due_days = int(default_due_days)
            
            # Handle logo upload / removal regardless of which tab is active.
            # This allows updating the logo and then adjusting another tab
            # (e.g. tax) before saving, without losing the logo change.
            if 'logo' in request.FILES:
                business.logo = request.FILES['logo']
            if request.POST.get('remove_logo') == '1':
                if business.logo:
                    business.logo.delete(save=False)
                business.logo = None

            business.save()
            messages.success(request, 'Settings saved successfully.')
            return redirect('accounting_invoice_settings')
        except Exception as e:
            messages.error(request, f'Error saving settings: {str(e)}')
    
    context = {
        'business': business,
        'templates': templates,
        'tax_slabs': tax_slabs,
    }
    return render(request, 'accounting/settings.html', context)


@login_required
def email_templates(request):
    """Manage email templates"""
    # Initialize default templates if they don't exist
    template_types = ['invoice_sent', 'invoice_reminder', 'payment_received']
    for template_type in template_types:
        EmailTemplate.objects.get_or_create(
            template_type=template_type,
            defaults={
                'subject': f'Invoice {template_type.replace("_", " ").title()}',
                'body': f'Default template for {template_type.replace("_", " ")}',
            }
        )
    
    templates = EmailTemplate.objects.all()
    
    if request.method == 'POST':
        template_id = request.POST.get('template_id')
        template = get_object_or_404(EmailTemplate, id=template_id)
        template.subject = request.POST.get('subject')
        template.body = request.POST.get('body')
        template.is_active = request.POST.get('is_active') == 'on'
        template.save()
        messages.success(request, 'Email template updated successfully.')
        return redirect('accounting_email_templates')
    
    context = {
        'templates': templates,
    }
    return render(request, 'accounting/email_templates.html', context)


@login_required
def invoice_preview(request, invoice_id):
    """Preview invoice"""
    today = timezone.now().date()
    business = BusinessInfo.get_current()
    
    if invoice_id == 0:
        # Preview for new invoice (draft) - check if POST data is provided
        if request.method == 'POST':
            # Build draft invoice from POST data
            draft_invoice = build_draft_invoice_from_post(request)
            context = {
                'invoice': draft_invoice,
                'business': business,
                'is_draft': True,
                'today': today,
            }
        else:
            # No POST data, show empty draft
            context = {
                'invoice': None,
                'business': business,
                'is_draft': True,
                'today': today,
            }
    else:
        invoice = get_object_or_404(Invoice, id=invoice_id)
        # For saved invoices, show quantity column only if any item quantity != 1
        # Add these attributes to the invoice object for template compatibility
        invoice.show_quantity = invoice.items.exclude(quantity=1).exists()
        invoice.show_price = True
        context = {
            'invoice': invoice,
            'business': business,
            'is_draft': False,
            'today': today,
        }
    
    return render(request, 'accounting/invoice_preview_standalone.html', context)


def build_draft_invoice_from_post(request):
    """Build a draft invoice object from POST data for preview"""
    from collections import namedtuple
    
    # Create a simple object to mimic Invoice model
    DraftInvoice = namedtuple('DraftInvoice', [
        'id',
        'title',
        'summary',
        'invoice_number',
        'invoice_date',
        'due_date',
        'po_number',
        'subtotal',
        'tax_amount',
        'discount_amount',
        'total_amount',
        'tax_rate',
        'notes',
        'footer',
        'customer',
        'applicant',
        'items',
        'show_quantity',
        'show_price',
    ])
    
    DraftItem = namedtuple('DraftItem', [
        'service', 'quantity', 'unit_price', 'description', 'line_total', 'is_taxable'
    ])
    
    # Legacy "customer" is no longer used in the new flow but we keep the
    # field on the draft object for backward compatibility. Applicant is
    # now the primary party on the invoice.
    customer = None

    # Get applicant from the hidden select field
    applicant = None
    applicant_id = request.POST.get('applicant')
    if applicant_id:
        try:
            applicant = Applicant.objects.get(id=applicant_id)
        except Applicant.DoesNotExist:
            applicant = None
    
    # Get form data
    title = request.POST.get('title', 'Invoice')
    summary = request.POST.get('summary', '')
    invoice_number = request.POST.get('invoice_number', '')
    invoice_date_str = request.POST.get('invoice_date', '')
    due_date_str = request.POST.get('due_date', '')
    po_number = request.POST.get('po_number', '')
    notes = request.POST.get('notes', '')
    footer = request.POST.get('footer', '')
    
    # Parse dates
    invoice_date = None
    due_date = None
    if invoice_date_str:
        try:
            invoice_date = datetime.strptime(invoice_date_str, '%Y-%m-%d').date()
        except:
            invoice_date = timezone.now().date()
    else:
        invoice_date = timezone.now().date()
    
    if due_date_str:
        try:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        except:
            due_date = invoice_date
    else:
        due_date = invoice_date
    
    # Get items from the new applicant-based invoice form
    service_names = request.POST.getlist('service_name[]')
    service_ids = request.POST.getlist('service[]')
    quantities = request.POST.getlist('quantity[]')
    unit_prices = request.POST.getlist('unit_price[]')
    is_taxable_flags = request.POST.getlist('is_taxable[]')
    
    items = []
    subtotal = Decimal('0.00')
    taxable_total = Decimal('0.00')

    max_len = max(
        len(service_names),
        len(service_ids),
        len(quantities),
        len(unit_prices),
        len(is_taxable_flags),
    ) if any([service_names, service_ids, quantities, unit_prices, is_taxable_flags]) else 0

    for i in range(max_len):
        name = service_names[i] if i < len(service_names) else ''
        service_id = service_ids[i] if i < len(service_ids) else ''
        qty_raw = quantities[i] if i < len(quantities) else ''
        price_raw = unit_prices[i] if i < len(unit_prices) else ''

        # Skip completely empty rows
        if not any([name, service_id, qty_raw, price_raw]):
            continue

        # Resolve optional rio.Services reference
        service_obj = None
        if service_id:
            try:
                service_obj = Services.objects.get(id=service_id)
            except Services.DoesNotExist:
                service_obj = None

        try:
            qty = Decimal(str(qty_raw)) if qty_raw else Decimal('1.00')
        except Exception:
            qty = Decimal('1.00')

        try:
            price = Decimal(str(price_raw)) if price_raw else Decimal('0.00')
        except Exception:
            price = Decimal('0.00')

        # Skip items with zero price and no name
        if price == 0 and not name:
            continue

        is_taxable = False
        if i < len(is_taxable_flags):
            is_taxable = str(is_taxable_flags[i]).strip() == '1'

        line_total = qty * price
        subtotal += line_total
        if is_taxable:
            taxable_total += line_total

        items.append(
            DraftItem(
                service=service_obj,
                quantity=qty,
                unit_price=price,
                description=name,
                line_total=line_total,
                is_taxable=is_taxable,
            )
        )
    
    # Calculate discount (applied BEFORE tax, with same rules as model/helper)
    discount_value = request.POST.get('discount_amount', '0') or '0'
    discount_type = request.POST.get('discount_type', 'amount')
    try:
        discount_decimal = Decimal(discount_value)
    except:
        discount_decimal = Decimal('0.00')
    
    discount_amount = Decimal('0.00')
    if discount_decimal > 0:
        if discount_type == 'percent':
            discount_amount = (subtotal * discount_decimal) / Decimal('100.00')
        else:
            discount_amount = discount_decimal
    
    if discount_amount > subtotal:
        discount_amount = subtotal

    # Effective discount ratio across all items
    discount_base = subtotal if subtotal > 0 else Decimal('0.00')
    if discount_base > 0:
        discount_ratio = (discount_amount / discount_base).quantize(
            Decimal('0.0001')
        )
    else:
        discount_ratio = Decimal('0.00')
    
    # Calculate tax on discounted taxable base only
    tax_rate = Decimal('0.00')
    tax_slab_id = request.POST.get('tax_slab') or ''
    if tax_slab_id:
        try:
            tax_slab = TaxSlab.objects.get(id=tax_slab_id, is_active=True)
            tax_rate = tax_slab.rate
        except TaxSlab.DoesNotExist:
            pass

    taxable_base = taxable_total * (Decimal('1.00') - discount_ratio)
    if taxable_base < 0:
        taxable_base = Decimal('0.00')

    tax_amount = Decimal('0.00')
    if tax_rate > 0 and taxable_base > 0:
        tax_amount = (taxable_base * tax_rate) / Decimal('100.00')

    discounted_subtotal = subtotal - discount_amount
    if discounted_subtotal < 0:
        discounted_subtotal = Decimal('0.00')

    total_amount = discounted_subtotal + tax_amount
    
    # Get column visibility settings
    show_quantity = request.POST.get('show_quantity', '1') == '1'
    show_price = request.POST.get('show_price', '1') == '1'
    
    # Create draft invoice object
    draft_invoice = DraftInvoice(
        id=None,  # Draft invoices don't have an ID
        title=title,
        summary=summary,
        invoice_number=invoice_number or 'DRAFT',
        invoice_date=invoice_date,
        due_date=due_date,
        po_number=po_number,
        subtotal=subtotal,
        tax_amount=tax_amount,
        discount_amount=discount_amount,
        total_amount=total_amount,
        tax_rate=tax_rate,
        notes=notes,
        footer=footer,
        customer=customer,
        applicant=applicant,
        items=items,
        show_quantity=show_quantity,
        show_price=show_price
    )
    
    return draft_invoice


# AJAX endpoints for DataTables server-side processing

@login_required
def datatable_transactions(request):
    """AJAX endpoint for transactions datatable"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get DataTables parameters
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))
    search_value = request.GET.get('search[value]', '')
    order_column = int(request.GET.get('order[0][column]', 1))
    order_dir = request.GET.get('order[0][dir]', 'desc')
    
    # Get filter parameters
    account_filter = request.GET.get('account', '')
    type_filter = request.GET.get('type', '')
    category_filter = request.GET.get('category', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    transactions = Transaction.objects.select_related('account', 'category').all()
    
    # Apply filters
    if account_filter:
        transactions = transactions.filter(account_id=account_filter)
    if type_filter:
        transactions = transactions.filter(transaction_type=type_filter)
    if category_filter:
        transactions = transactions.filter(category_id=category_filter)
    if date_from:
        transactions = transactions.filter(transaction_date__gte=date_from)
    if date_to:
        transactions = transactions.filter(transaction_date__lte=date_to)
    
    # Apply search
    if search_value:
        transactions = transactions.filter(
            Q(description__icontains=search_value) |
            Q(account__name__icontains=search_value) |
            Q(category__name__icontains=search_value) |
            Q(reference_number__icontains=search_value)
        )
    
    # Get total count (before search/filtering)
    records_total = Transaction.objects.count()
    # Get filtered count (after search/filtering)
    records_filtered = transactions.count()
    
    # Ordering
    order_columns = ['id', 'transaction_date', 'description', 'account__name', 'category__name', 'amount', 'id']
    order_by = order_columns[order_column] if order_column < len(order_columns) else 'transaction_date'
    if order_dir == 'desc':
        order_by = '-' + order_by
    transactions = transactions.order_by(order_by)
    
    # Pagination
    paginator = Paginator(transactions, length)
    page = (start // length) + 1
    page_obj = paginator.get_page(page)
    
    # Format data
    data = []
    for transaction in page_obj:
        data.append([
            f'<input type="checkbox" class="transaction-checkbox" value="{transaction.id}">',
            transaction.transaction_date.strftime('%b %d, %Y'),
            transaction.description[:50] + '...' if len(transaction.description) > 50 else transaction.description,
            transaction.account.name,
            transaction.category.name if transaction.category else '<span style="color: #6c757d;">Uncategorized</span>',
            f'<span class="{"amount-positive" if transaction.transaction_type == "deposit" else "amount-negative"}">{"+" if transaction.transaction_type == "deposit" else "-"}${transaction.amount:.2f}</span>',
            f'<div class="hstack gap-2 justify-content-start">'
            f'<a href="#" class="avatar-text avatar-md edit-transaction-btn" data-transaction-id="{transaction.id}" title="Edit Transaction"><i class="feather feather-edit-2"></i></a>'
            f'<a href="#" class="avatar-text avatar-md delete-transaction-btn" data-transaction-id="{transaction.id}" title="Delete Transaction"><i class="feather feather-trash-2"></i></a>'
            f'</div>'
        ])
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })


@login_required
def get_applicant_invoices(request, applicant_id):
    """
    AJAX endpoint to get an applicant's existing invoices.

    Used to:
    - Warn about potential duplicate billing on the invoice create page
    - Optionally preselect an existing invoice elsewhere in the UI
    """
    applicant = get_object_or_404(Applicant, id=applicant_id)

    # Consider all non-cancelled invoices as "existing" for duplicate checks.
    invoices = (
        Invoice.objects.filter(applicant=applicant)
        .exclude(status='cancelled')
        .order_by('-invoice_date', '-id')
    )

    return JsonResponse(
        {
            'invoices': [
                {
                    'id': inv.id,
                    'invoice_number': inv.invoice_number,
                    'status': inv.status,
                    'status_display': inv.get_status_display(),
                    'invoice_date': inv.invoice_date.isoformat() if inv.invoice_date else None,
                    'total_amount': str(inv.total_amount),
                    'balance': str(inv.get_balance()),
                    'url': reverse('accounting_invoice_view', args=[inv.id]),
                }
                for inv in invoices
            ]
        }
    )


@login_required
def get_applicant_data(request, applicant_id):
    """
    AJAX endpoint used by the invoice create page to auto-populate items
    from an applicant's configured fee data (application fee, biometric
    fee, client retained amount, etc.) and to show a small payment
    summary.
    """
    applicant = get_object_or_404(Applicant, id=applicant_id)

    # Build base applicant info
    applicant_data = {
        'id': applicant.id,
        'client_id': applicant.client_id,
        'f_name': applicant.f_name,
        'm_name': applicant.m_name,
        'l_name': applicant.l_name,
        'email': applicant.email,
        'phone': applicant.phone,
        'address': applicant.address,
        'application_type': applicant.application_type,
        'application_fee_amount': float(applicant.application_fee_amount or 0) if applicant.application_fee_amount is not None else 0.0,
        'biometric_fee_amount': float(applicant.biometric_fee_amount or 0) if applicant.biometric_fee_amount is not None else 0.0,
        'client_retained_amount': float(applicant.client_retained_amount or 0) if applicant.client_retained_amount is not None else 0.0,
        'total_fee_amount': float(applicant.total_fee_amount or 0) if applicant.total_fee_amount is not None else 0.0,
        'balance_amount': float(applicant.balance_amount or 0) if applicant.balance_amount is not None else 0.0,
    }

    # Suggested invoice items for the UI auto-populate helper.
    # Service fee should always appear FIRST, then application fee, then biometric.
    suggested_items = []

    # Client retained amount / service fee (non-taxable) – always on top
    if applicant.client_retained_amount:
        suggested_items.append(
            {
                'name': f"Service fee for {applicant.application_type or 'service'}",
                'quantity': 1,
                'unit_price': float(applicant.client_retained_amount),
                'type': 'service_fee',
            }
        )

    # Application fee (taxable)
    if applicant.application_fee_amount:
        suggested_items.append(
            {
                'name': f"Application fee for {applicant.application_type or 'service'}",
                'quantity': 1,
                'unit_price': float(applicant.application_fee_amount),
                'type': 'application_fee',
            }
        )

    # Biometric fee (non-taxable)
    if applicant.biometric_fee_amount:
        suggested_items.append(
            {
                'name': 'Biometric fee',
                'quantity': 1,
                'unit_price': float(applicant.biometric_fee_amount),
                'type': 'biometric_fee',
            }
        )

    # Fallback to a single line from total fee if there were no granular items
    if not suggested_items and applicant.total_fee_amount:
        suggested_items.append(
            {
                'name': f"Total service fee for {applicant.application_type or 'service'}",
                'quantity': 1,
                'unit_price': float(applicant.total_fee_amount),
                'type': 'service_total',
            }
        )

    # Payment history from rio.Payments
    applicant_payments = ApplicantPayment.objects.filter(
        applicant=applicant
    ).order_by('-paid_date', '-id')

    payment_history = []
    total_paid = 0.0
    for pay in applicant_payments:
        amount_val = float(pay.amount or 0)
        total_paid += amount_val
        payment_history.append(
            {
                'amount': amount_val,
                'payment_method': getattr(pay, 'payment_method', None),
                'paid_date': pay.paid_date.strftime('%Y-%m-%d') if pay.paid_date else None,
                'added_at': pay.added_at.strftime('%Y-%m-%d') if pay.added_at else None,
            }
        )

    # Service info summary for the side panel
    service_info = {
        'total_fee_amount': applicant_data['total_fee_amount'],
        'balance_amount': applicant_data['balance_amount'],
    }

    return JsonResponse(
        {
            'has_applicant': True,
            'applicant': applicant_data,
            'service_info': service_info,
            'payment_history': payment_history,
            'total_paid': total_paid,
            'suggested_items': suggested_items,
        }
    )


@login_required
def datatable_customers(request):
    """AJAX endpoint for customers datatable"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))
    search_value = request.GET.get('search[value]', '')
    order_column = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'asc')
    
    customers = Customer.objects.all()
    
    # Apply search
    if search_value:
        customers = customers.filter(
            Q(name__icontains=search_value) |
            Q(email__icontains=search_value) |
            Q(phone__icontains=search_value)
        )
    
    records_total = Customer.objects.count()
    records_filtered = customers.count()
    
    # Ordering
    order_columns = ['name', 'email', 'phone', 'city', 'id', 'id', 'id', 'is_active', 'id']
    order_by = order_columns[order_column] if order_column < len(order_columns) else 'name'
    if order_dir == 'desc':
        order_by = '-' + order_by
    customers = customers.order_by(order_by)
    
    paginator = Paginator(customers, length)
    page = (start // length) + 1
    page_obj = paginator.get_page(page)
    
    data = []
    for customer in page_obj:
        data.append([
            customer.name,
            customer.email or '-',
            customer.phone or '-',
            customer.city or '-',
            f'${customer.get_total_invoiced():.2f}',
            f'${customer.get_total_paid():.2f}',
            f'${customer.get_balance():.2f}',
            f'<span class="badge bg-{"success" if customer.is_active else "secondary"}">{"Active" if customer.is_active else "Inactive"}</span>',
            f'<a href="{reverse("accounting_customer_edit", args=[customer.id])}" class="btn btn-sm btn-warning">Edit</a>'
        ])
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })


@login_required
def datatable_expenses(request):
    """AJAX endpoint for expenses datatable"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))
    search_value = request.GET.get('search[value]', '')
    order_column = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'desc')
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    expenses = Expense.objects.select_related('category').all()
    
    # Apply filters
    if status_filter:
        expenses = expenses.filter(payment_status=status_filter)
    if date_from:
        expenses = expenses.filter(expense_date__gte=date_from)
    if date_to:
        expenses = expenses.filter(expense_date__lte=date_to)
    
    # Apply search
    if search_value:
        expenses = expenses.filter(
            Q(vendor__icontains=search_value) |
            Q(description__icontains=search_value) |
            Q(reference_number__icontains=search_value)
        )
    
    records_total = Expense.objects.count()
    records_filtered = expenses.count()
    
    # Ordering
    order_columns = ['expense_date', 'vendor', 'description', 'amount', 'category__name', 'payment_status', 'id']
    order_by = order_columns[order_column] if order_column < len(order_columns) else 'expense_date'
    if order_dir == 'desc':
        order_by = '-' + order_by
    expenses = expenses.order_by(order_by)
    
    paginator = Paginator(expenses, length)
    page = (start // length) + 1
    page_obj = paginator.get_page(page)
    
    data = []
    for expense in page_obj:
        status_badge_class = 'success' if expense.payment_status == 'paid' else ('warning' if expense.payment_status == 'partial' else 'danger')
        receipt_btn = f'<a href="{expense.receipt.url}" target="_blank" class="btn btn-sm btn-info">View Receipt</a>' if expense.receipt else ''
        data.append([
            expense.expense_date.strftime('%b %d, %Y'),
            expense.vendor,
            expense.description[:50] + '...' if len(expense.description) > 50 else expense.description,
            f'${expense.amount:.2f}',
            expense.category.name if expense.category else '-',
            f'<span class="badge bg-{status_badge_class}">{expense.get_payment_status_display()}</span>',
            receipt_btn
        ])
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })


@login_required
def datatable_products(request):
    """AJAX endpoint for products/services datatable"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))
    search_value = request.GET.get('search[value]', '')
    order_column = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'asc')
    
    # Get filter parameters
    type_filter = request.GET.get('type', '')
    
    products = ProductService.objects.all()
    
    # Apply filters
    if type_filter:
        products = products.filter(type=type_filter)
    
    # Apply search
    if search_value:
        products = products.filter(
            Q(name__icontains=search_value) |
            Q(sku__icontains=search_value) |
            Q(description__icontains=search_value)
        )
    
    records_total = ProductService.objects.count()
    records_filtered = products.count()
    
    # Ordering
    order_columns = ['name', 'type', 'sku', 'price', 'cost', 'tax_rate', 'is_active', 'id']
    order_by = order_columns[order_column] if order_column < len(order_columns) else 'name'
    if order_dir == 'desc':
        order_by = '-' + order_by
    products = products.order_by(order_by)
    
    paginator = Paginator(products, length)
    page = (start // length) + 1
    page_obj = paginator.get_page(page)
    
    data = []
    for product in page_obj:
        type_badge_class = 'info' if product.type == 'product' else 'primary'
        data.append([
            product.name,
            f'<span class="badge bg-{type_badge_class}">{product.get_type_display()}</span>',
            product.sku or '-',
            f'${product.price:.2f}',
            f'${product.cost:.2f}',
            f'{product.tax_rate:.2f}%',
            f'<span class="badge bg-{"success" if product.is_active else "secondary"}">{"Active" if product.is_active else "Inactive"}</span>',
            f'<a href="{reverse("accounting_product_service_edit", args=[product.id])}" class="btn btn-sm btn-warning">Edit</a>'
        ])
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })


@login_required
def datatable_invoices(request):
    """AJAX endpoint for invoices datatable"""
    from django.core.paginator import Paginator
    from django.db.models import Q, Sum
    from django.db.models.functions import Coalesce
    from django.db.models import Value, DecimalField

    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))
    search_value = request.GET.get('search[value]', '')
    order_column = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'desc')

    invoice_number_filter = request.GET.get('invoice_number', '')
    status_filter = request.GET.get('status', '')
    customer_filter = request.GET.get('customer', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    invoices = Invoice.objects.select_related('applicant').all()

    if invoice_number_filter:
        invoices = invoices.filter(invoice_number__icontains=invoice_number_filter)
    if status_filter:
        invoices = invoices.filter(status=status_filter)
    if customer_filter:
        invoices = invoices.filter(applicant_id=customer_filter)
    if date_from:
        invoices = invoices.filter(invoice_date__gte=date_from)
    if date_to:
        invoices = invoices.filter(invoice_date__lte=date_to)

    if search_value:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search_value) |
            Q(applicant__f_name__icontains=search_value) |
            Q(applicant__l_name__icontains=search_value) |
            Q(applicant__client_id__icontains=search_value)
        )

    records_total = Invoice.objects.count()
    records_filtered = invoices.count()

    order_columns = [
        'invoice_number', 'applicant__f_name', 'invoice_date',
        'due_date', 'total_amount', 'id', 'id', 'status', 'id',
    ]
    order_by = order_columns[order_column] if order_column < len(order_columns) else 'invoice_number'
    if order_dir == 'desc':
        order_by = '-' + order_by
    invoices = invoices.order_by(order_by)

    paginator = Paginator(invoices, length)
    page = (start // length) + 1
    page_obj = paginator.get_page(page)

    # ─── FIX: Pre-fetch rio payments for all invoices on this page ───
    # Collect applicant IDs for this page only (efficient, no N+1)
    page_invoices = list(page_obj)
    applicant_ids = [
        inv.applicant_id for inv in page_invoices if inv.applicant_id
    ]

    # Sum rio payments per applicant in ONE query
    rio_paid_map = dict(
        ApplicantPayment.objects.filter(applicant_id__in=applicant_ids)
        .values('applicant_id')
        .annotate(total=Sum('amount'))
        .values_list('applicant_id', 'total')
    )

    # Sum accounting payments per invoice in ONE query
    accounting_paid_map = dict(
        Payment.objects.filter(invoice__in=page_invoices)
        .values('invoice_id')
        .annotate(total=Sum('amount'))
        .values_list('invoice_id', 'total')
    )

    data = []
    for invoice in page_invoices:
        applicant = getattr(invoice, 'applicant', None)
        if applicant is not None:
            name_parts = [
                getattr(applicant, 'f_name', '') or '',
                getattr(applicant, 'l_name', '') or '',
            ]
            applicant_name = ' '.join(p for p in name_parts if p).strip() or applicant.client_id
        else:
            applicant_name = '-'

        # Use rio payments (same source as detail page) with accounting as fallback
        rio_paid = Decimal(str(rio_paid_map.get(invoice.applicant_id) or '0.00'))
        accounting_paid = Decimal(str(accounting_paid_map.get(invoice.id) or '0.00'))

        # Mirror view_invoice logic: prefer rio payments if they exist
        paid_amount = rio_paid if rio_paid > Decimal('0.00') else accounting_paid
        balance = invoice.total_amount - paid_amount
        if balance < Decimal('0.00'):
            balance = Decimal('0.00')

        status_class = f'status-{invoice.status}'

        actions_html = '<div class="hstack gap-2 justify-content-start">'
        actions_html += f'<a href="{reverse("accounting_invoice_view", args=[invoice.id])}" title="View Invoice" class="avatar-text avatar-md"><i class="feather feather-eye"></i></a>'
        if not invoice.payments.exists():
            actions_html += f'<a href="{reverse("accounting_invoice_edit", args=[invoice.id])}" title="Edit Invoice" class="avatar-text avatar-md"><i class="feather feather-edit-2"></i></a>'
        actions_html += f'<a href="{reverse("accounting_invoice_preview", args=[invoice.id])}" target="_blank" title="Preview Invoice" class="avatar-text avatar-md"><i class="feather feather-file-text"></i></a>'
        payment_url = f'{reverse("accounting_payment_create")}?invoice={invoice.id}'
        actions_html += f'<a href="{payment_url}" title="Record Payment" class="avatar-text avatar-md"><i class="feather feather-dollar-sign"></i></a>'
        actions_html += f'<a href="{reverse("accounting_invoice_preview", args=[invoice.id])}" target="_blank" onclick="window.open(this.href).print(); return false;" title="Download/Print Invoice" class="avatar-text avatar-md"><i class="feather feather-download"></i></a>'
        actions_html += '</div>'

        data.append([
            invoice.invoice_number,
            applicant_name,
            invoice.invoice_date.strftime('%b %d, %Y'),
            invoice.due_date.strftime('%b %d, %Y'),
            f'${invoice.total_amount:.2f}',
            f'${paid_amount:.2f}',       # ✅ Correct paid amount
            f'${balance:.2f}',           # ✅ Correct balance
            f'<span class="status-badge {status_class}">{invoice.get_status_display()}</span>',
            actions_html,
        ])

    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data,
    })

@login_required
def datatable_payments(request):
    """AJAX endpoint for payments datatable"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))
    search_value = request.GET.get('search[value]', '')
    order_column = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'desc')
    
    # Get filter parameters ("customer" now represents an Applicant ID)
    customer_filter = request.GET.get('customer', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    payments = Payment.objects.select_related('applicant', 'invoice').all()
    
    # Apply filters
    if customer_filter:
        payments = payments.filter(applicant_id=customer_filter)
    if date_from:
        payments = payments.filter(payment_date__gte=date_from)
    if date_to:
        payments = payments.filter(payment_date__lte=date_to)
    
    # Apply search
    if search_value:
        payments = payments.filter(
            Q(applicant__f_name__icontains=search_value) |
            Q(applicant__l_name__icontains=search_value) |
            Q(applicant__client_id__icontains=search_value) |
            Q(invoice__invoice_number__icontains=search_value) |
            Q(reference_number__icontains=search_value)
        )
    
    records_total = Payment.objects.count()
    records_filtered = payments.count()
    
    # Ordering
    order_columns = [
        'payment_date',
        'applicant__f_name',
        'invoice__invoice_number',
        'amount',
        'payment_method',
        'reference_number',
    ]
    order_by = order_columns[order_column] if order_column < len(order_columns) else 'payment_date'
    if order_dir == 'desc':
        order_by = '-' + order_by
    payments = payments.order_by(order_by)
    
    paginator = Paginator(payments, length)
    page = (start // length) + 1
    page_obj = paginator.get_page(page)
    
    data = []
    for payment in page_obj:
        applicant = getattr(payment, 'applicant', None)
        if applicant is not None:
            name_parts = [
                getattr(applicant, 'f_name', '') or '',
                getattr(applicant, 'l_name', '') or '',
            ]
            applicant_name = ' '.join(p for p in name_parts if p).strip() or applicant.client_id
        else:
            applicant_name = '-'

        data.append(
            [
                payment.payment_date.strftime('%b %d, %Y'),
                applicant_name,
                payment.invoice.invoice_number if payment.invoice else '-',
                f'<span class="text-success">${payment.amount:.2f}</span>',
                payment.get_payment_method_display(),
                payment.reference_number or '-',
            ]
        )
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })